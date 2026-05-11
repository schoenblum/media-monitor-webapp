"""Outlet CRUD plus the XLSX bulk import / export endpoints."""
import io
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_user
from app.models.outlet import Outlet
from app.models.user import User
from app.schemas.outlet import (
    ImportReport,
    ImportReportRow,
    OutletCreate,
    OutletOut,
    OutletUpdate,
)
from app.schemas.outlet import _clean_domain  # type: ignore[attr-defined]
from app.services.default_outlets import SUPPORTED_LANGUAGES


router = APIRouter(prefix="/outlets", tags=["outlets"])

TEMPLATE_HEADERS = ("name", "domain", "category", "keyword_langs", "notes")


def _outlet_to_out(o: Outlet) -> OutletOut:
    return OutletOut.model_validate(o)


@router.get("", response_model=list[OutletOut])
async def list_outlets(
    category: str | None = None,
    active: bool | None = None,
    current: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[OutletOut]:
    stmt = select(Outlet).where(Outlet.user_id == current.id)
    if category:
        stmt = stmt.where(Outlet.category == category)
    if active is not None:
        stmt = stmt.where(Outlet.is_active == active)
    stmt = stmt.order_by(Outlet.name.asc())
    rows = (await db.execute(stmt)).scalars().all()
    return [_outlet_to_out(o) for o in rows]


@router.post("", response_model=OutletOut, status_code=status.HTTP_201_CREATED)
async def create_outlet(
    payload: OutletCreate,
    current: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> OutletOut:
    existing = (
        await db.execute(
            select(Outlet).where(Outlet.user_id == current.id, Outlet.domain == payload.domain)
        )
    ).scalar_one_or_none()
    if existing is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT, detail="An outlet with that domain already exists"
        )
    o = Outlet(
        user_id=current.id,
        name=payload.name.strip(),
        domain=payload.domain,
        category=payload.category,
        keyword_langs=payload.keyword_langs,
        is_active=payload.is_active,
    )
    db.add(o)
    await db.commit()
    await db.refresh(o)
    return _outlet_to_out(o)


@router.put("/{outlet_id}", response_model=OutletOut)
async def update_outlet(
    outlet_id: UUID,
    payload: OutletUpdate,
    current: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> OutletOut:
    o = await db.get(Outlet, outlet_id)
    if o is None or o.user_id != current.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outlet not found")
    data = payload.model_dump(exclude_unset=True)
    if "domain" in data and data["domain"]:
        if data["domain"] != o.domain:
            clash = (
                await db.execute(
                    select(Outlet).where(
                        Outlet.user_id == current.id, Outlet.domain == data["domain"]
                    )
                )
            ).scalar_one_or_none()
            if clash is not None and clash.id != o.id:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT, detail="Domain already used by another outlet"
                )
    for key, value in data.items():
        setattr(o, key, value)
    await db.commit()
    await db.refresh(o)
    return _outlet_to_out(o)


@router.delete("/{outlet_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_outlet(
    outlet_id: UUID,
    current: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> None:
    o = await db.get(Outlet, outlet_id)
    if o is None or o.user_id != current.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Outlet not found")
    await db.delete(o)
    await db.commit()


# ---------------------------------------------------------------------------
# Import / export
# ---------------------------------------------------------------------------


def _build_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "outlets"
    ws.append(list(TEMPLATE_HEADERS))
    # Add one example row so users see the expected shape.
    ws.append([
        "Example Outlet",
        "example.com",
        "Outstanding international importance",
        "en,de",
        "comma-separated language codes; notes column is ignored on import",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/import/template")
async def download_import_template(_: User = Depends(get_current_user)) -> StreamingResponse:
    data = _build_template_bytes()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="outlet_import_template.xlsx"'},
    )


@router.get("/export")
async def export_outlets(
    current: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)
) -> StreamingResponse:
    rows = (
        await db.execute(
            select(Outlet).where(Outlet.user_id == current.id).order_by(Outlet.name.asc())
        )
    ).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "outlets"
    ws.append(list(TEMPLATE_HEADERS))
    for o in rows:
        ws.append([
            o.name,
            o.domain,
            o.category or "",
            ",".join(o.keyword_langs or []),
            "active" if o.is_active else "inactive",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="outlets_export.xlsx"'},
    )


@router.post("/import", response_model=ImportReport)
async def import_outlets(
    file: UploadFile = File(...),
    mode: str = Query("add", pattern="^(replace|add)$"),
    current: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ImportReport:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload an .xlsx file"
        )
    body = await file.read()
    try:
        wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=f"Could not parse XLSX: {exc}"
        ) from exc
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty workbook")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No header row")
    header_map = {str(h).strip().lower() if h else "": i for i, h in enumerate(headers)}
    for required in ("name", "domain"):
        if required not in header_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required column: {required}",
            )

    if mode == "replace":
        existing_rows = (
            await db.execute(select(Outlet).where(Outlet.user_id == current.id))
        ).scalars().all()
        for o in existing_rows:
            await db.delete(o)
        await db.flush()
        existing_domains: set[str] = set()
    else:
        existing_domains = {
            d for (d,) in (
                await db.execute(select(Outlet.domain).where(Outlet.user_id == current.id))
            ).all()
        }

    imported = 0
    skipped: list[ImportReportRow] = []

    for i, row in enumerate(rows_iter, start=2):  # row 1 is the header
        try:
            name = str(row[header_map["name"]]).strip() if row[header_map["name"]] else ""
            domain_raw = str(row[header_map["domain"]]).strip() if row[header_map["domain"]] else ""
        except IndexError:
            skipped.append(ImportReportRow(row=i, reason="Row too short"))
            continue
        if not name or not domain_raw:
            skipped.append(ImportReportRow(row=i, reason="Missing name or domain"))
            continue
        domain = _clean_domain(domain_raw)
        if not domain:
            skipped.append(ImportReportRow(row=i, reason="Invalid domain"))
            continue
        if domain in existing_domains:
            skipped.append(ImportReportRow(row=i, reason=f"Duplicate domain: {domain}"))
            continue

        category = None
        if "category" in header_map and row[header_map["category"]]:
            category = str(row[header_map["category"]]).strip() or None
        langs: list[str] = []
        if "keyword_langs" in header_map and row[header_map["keyword_langs"]]:
            raw_langs = str(row[header_map["keyword_langs"]])
            lang_error = False
            for token in raw_langs.replace(";", ",").split(","):
                t = token.strip().lower()
                if not t:
                    continue
                if t not in SUPPORTED_LANGUAGES:
                    skipped.append(ImportReportRow(row=i, reason=f"Unsupported language code: {t}"))
                    lang_error = True
                    break
                langs.append(t)
            if lang_error:
                continue
        if not langs:
            langs = ["en"]

        db.add(
            Outlet(
                user_id=current.id,
                name=name,
                domain=domain,
                category=category,
                keyword_langs=langs,
                is_active=True,
            )
        )
        existing_domains.add(domain)
        imported += 1

    await db.commit()
    return ImportReport(imported=imported, skipped=skipped)
