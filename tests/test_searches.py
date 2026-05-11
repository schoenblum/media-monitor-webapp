"""Search CRUD + outlet linking + import template / export round-trip."""
import io

import pytest
from openpyxl import load_workbook


async def _login(client):
    r = await client.post(
        "/api/v1/auth/login",
        json={"email": "admin@test.local", "password": "TestAdminPassword!"},
    )
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


@pytest.mark.asyncio
async def test_bootstrap_seeded_outlets_and_default_search(client):
    h = await _login(client)
    r = await client.get("/api/v1/outlets", headers=h)
    assert r.status_code == 200
    assert len(r.json()) > 50  # the seeded library is ~75 outlets

    r = await client.get("/api/v1/searches", headers=h)
    assert r.status_code == 200
    searches = r.json()
    assert any(s["is_default"] for s in searches)


@pytest.mark.asyncio
async def test_outlet_create_and_duplicate_domain_rejected(client):
    h = await _login(client)
    payload = {
        "name": "Test Outlet",
        "domain": "https://example.test/",
        "category": "Custom",
        "keyword_langs": ["en", "fr"],
        "is_active": True,
    }
    r = await client.post("/api/v1/outlets", json=payload, headers=h)
    assert r.status_code == 201
    body = r.json()
    assert body["domain"] == "example.test"  # cleaned

    # duplicate domain → 409
    r2 = await client.post("/api/v1/outlets", json=payload, headers=h)
    assert r2.status_code == 409


@pytest.mark.asyncio
async def test_outlet_import_template_and_export(client):
    h = await _login(client)
    r = await client.get("/api/v1/outlets/import/template", headers=h)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    headers = [c.value for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
    assert headers == ["name", "domain", "category", "keyword_langs", "notes"]

    r = await client.get("/api/v1/outlets/export", headers=h)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == ("name", "domain", "category", "keyword_langs", "notes")
    assert len(rows) > 1  # seeded outlets present
